"""
Deterministic Excel workbook generator for EWA analysis results.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ewa_pipeline.report.schemas import AnalysisResult, CrossReference, DomainAnalysis, Finding
from ewa_pipeline.tracking.token_tracker import TokenUsage

# ── Colours ───────────────────────────────────────────────────────────────────

_F_HDR  = "Cambria"   # header font
_F_BODY = "Calibri"   # body font
_SZ     = 11          # body size

_C_HDR_BG   = "4472C4"
_C_HDR_FG   = "FFFFFF"
_C_ALT      = "F2F2F2"
_C_LINK     = "4472C4"
_C_TITLE_BG = "1F4E79"
_C_BORDER   = "BFBFBF"

_SEV_BG = {"Critical": "C00000", "High": "ED7D31", "Medium": "FFC000", "Low": "548235"}
_SEV_FG = {"Critical": "FFFFFF", "High": "FFFFFF", "Medium": "000000", "Low": "FFFFFF"}
_HEALTH_BG = {"Critical": "C00000", "Warning": "ED7D31", "Healthy": "548235"}

# Tab colours
_TAB_EXEC = _C_TITLE_BG
_TAB_XREF = "7030A0"
_TAB_REM  = "2E75B6"
_TAB_UTIL = "808080"

_NUM = '#,##0'

# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, size=_SZ, color="000000", italic=False):
    return Font(name=_F_BODY, bold=bold, size=size, color=color, italic=italic)

def _hdr_font(color=_C_HDR_FG, size=_SZ):
    return Font(name=_F_HDR, bold=True, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_bottom(color=_C_BORDER):
    return Border(bottom=Side(style="thin", color=color))

def _apply_app_style(ws, *, hide_headers=False, zoom=110):
    """Hide gridlines (and optionally row/col headers) and set zoom."""
    ws.sheet_view.showGridLines = False
    if hide_headers:
        ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.zoomScale = zoom

def _hdr_row(ws, row, values, bg=_C_HDR_BG, fg=_C_HDR_FG, start_col=1):
    for i, val in enumerate(values):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=val)
        c.font = _hdr_font(color=fg)
        c.fill = _fill(bg)
        c.alignment = _align(h="center")
        c.border = _border_bottom()

def _sheet_ref(name: str) -> str:
    import re as _re
    return f"'{name}'" if _re.search(r"[^A-Za-z0-9_]", name) else name

def _autofit_columns(ws, min_width=8, max_width=55, skip_cols=()):
    """Set column widths based on longest content in each column."""
    col_widths: dict[int, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.column in skip_cols:
                continue
            col = cell.column
            length = min(len(str(cell.value)), max_width)
            if length > col_widths.get(col, 0):
                col_widths[col] = length
    for col, length in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, length + 2)

def _kpi_tile(ws, top_left, bottom_right, label, value, bg, fg="FFFFFF"):
    """Card-style KPI tile across a merged cell range."""
    ws.merge_cells(f"{top_left}:{bottom_right}")
    c = ws[top_left]
    c.value = f"{label}\n{value}"
    c.fill = _fill(bg)
    c.font = Font(name=_F_HDR, bold=True, size=18, color=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_executive_summary(wb: Workbook, result: AnalysisResult, sheet_map: dict[str, str]):
    ws = wb.create_sheet("Executive Summary")
    _apply_app_style(ws, hide_headers=True, zoom=110)
    ws.sheet_properties.tabColor = _TAB_EXEC

    # Margin columns A and I
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["I"].width = 2

    # Aggregate KPI counts
    totals = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for da in result.domain_analyses:
        for f in da.findings:
            if f.severity in totals:
                totals[f.severity] += 1
    total_findings = sum(totals.values())

    # Title band
    ws.row_dimensions[1].height = 40
    ws.merge_cells("B1:H1")
    c = ws["B1"]
    c.value = "SAP EarlyWatch Alert — Deep Analysis Report"
    c.font = Font(name=_F_HDR, bold=True, size=18, color="FFFFFF")
    c.fill = _fill(_C_TITLE_BG)
    c.alignment = _align(h="center")

    # Health banner
    health = result.overall_system_health
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = f"Overall System Health:  {str(health).upper()}"
    c.font = Font(name=_F_HDR, bold=True, size=13, color="FFFFFF")
    c.fill = _fill(_HEALTH_BG.get(health, "888888"))
    c.alignment = _align(h="center")

    # Spacer
    ws.row_dimensions[3].height = 10

    # KPI tile band — rows 4-6
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 12

    _kpi_tile(ws, "B4", "C6", "TOTAL FINDINGS", total_findings, _C_TITLE_BG)
    _kpi_tile(ws, "D4", "D6", "CRITICAL",       totals["Critical"], _SEV_BG["Critical"])
    _kpi_tile(ws, "E4", "E6", "HIGH",           totals["High"],     _SEV_BG["High"])
    _kpi_tile(ws, "F4", "F6", "MEDIUM",         totals["Medium"],   _SEV_BG["Medium"], fg="000000")
    _kpi_tile(ws, "G4", "G6", "LOW",            totals["Low"],      _SEV_BG["Low"])

    # Spacer
    ws.row_dimensions[7].height = 10

    # Top 5 Priority Actions header
    r = 8
    c = ws.cell(row=r, column=2, value="Top 5 Priority Actions")
    c.font = Font(name=_F_HDR, bold=True, size=13, color=_C_TITLE_BG)
    c.border = _border_bottom(_C_TITLE_BG)
    ws.merge_cells(f"B{r}:H{r}")
    for col in range(2, 9):
        ws.cell(row=r, column=col).border = _border_bottom(_C_TITLE_BG)
    ws.row_dimensions[r].height = 22

    r += 1
    for i, action in enumerate(result.top_5_priority_actions, 1):
        ws.cell(row=r, column=2, value=f"{i}").font = _font(bold=True, color=_C_TITLE_BG, size=12)
        ws.cell(row=r, column=2).alignment = _align(h="center", v="top")
        c = ws.cell(row=r, column=3, value=action)
        c.font = _font()
        c.alignment = _align(wrap=True, v="top")
        ws.merge_cells(f"C{r}:H{r}")
        if i % 2 == 1:
            for col in range(2, 9):
                ws.cell(row=r, column=col).fill = _fill(_C_ALT)
        ws.row_dimensions[r].height = 50
        r += 1

    # Spacer
    ws.row_dimensions[r].height = 10
    r += 1

    # Section table
    tbl_hdr_row = r
    col_headers = ["Section", "Health", "Findings", "Critical", "High", "Medium", "Low"]
    _hdr_row(ws, tbl_hdr_row, col_headers, start_col=2)
    ws.row_dimensions[tbl_hdr_row].height = 24
    ws.auto_filter.ref = f"B{tbl_hdr_row}:H{tbl_hdr_row}"

    r = tbl_hdr_row + 1
    for idx, da in enumerate(result.domain_analyses):
        if da.section_id not in sheet_map:
            continue
        sname = sheet_map[da.section_id]
        ref = _sheet_ref(sname)

        lc = ws.cell(row=r, column=2, value=da.section_title)
        lc.hyperlink = f"#'{sname}'!A1"
        lc.font = _font(color=_C_LINK)

        hc = ws.cell(row=r, column=3, value=da.overall_health)
        hc.font = _font(bold=True, color="FFFFFF")
        hc.alignment = _align(h="center")

        sev_col = "B"
        data_range = f"{ref}!${sev_col}$3:${sev_col}$9999"
        ws.cell(row=r, column=4, value=f"=COUNTA({ref}!$A$3:$A$9999)")
        ws.cell(row=r, column=5, value=f'=COUNTIF({data_range},"Critical")')
        ws.cell(row=r, column=6, value=f'=COUNTIF({data_range},"High")')
        ws.cell(row=r, column=7, value=f'=COUNTIF({data_range},"Medium")')
        ws.cell(row=r, column=8, value=f'=COUNTIF({data_range},"Low")')

        for col in range(4, 9):
            ws.cell(row=r, column=col).alignment = _align(h="center")
            ws.cell(row=r, column=col).font = _font()

        if idx % 2 == 0:
            for col in range(2, 9):
                ws.cell(row=r, column=col).fill = _fill(_C_ALT)
        # Health colour overrides alt shade
        hc.fill = _fill(_HEALTH_BG.get(da.overall_health, "888888"))

        ws.row_dimensions[r].height = 22
        r += 1

    _autofit_columns(ws, min_width=8, max_width=45, skip_cols=(1, 9))


def _build_section_sheet(wb: Workbook, da: DomainAnalysis, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    _apply_app_style(ws, zoom=110)
    ws.sheet_properties.tabColor = _HEALTH_BG.get(da.overall_health, _TAB_UTIL)

    COLS = ["ID", "Severity", "Title", "Description", "Evidence",
            "Impact", "Action", "Transactions", "Effort", "Priority",
            "External Validation", "Validation Reason", "Validation Query"]

    # Row 1: full-width back-link
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    bl = ws.cell(row=1, column=1, value="← Executive Summary")
    bl.hyperlink = "#'Executive Summary'!A1"
    bl.font = _font(color=_C_LINK)
    bl.alignment = _align(h="left")

    # Row 2: column headers
    _hdr_row(ws, 2, COLS)
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    for i, f in enumerate(da.findings):
        r = 3 + i
        vals = [
            f.id, f.severity, f.title, f.description, f.evidence,
            f.impact, f.remediation.action,
            ", ".join(f.remediation.sap_transactions),
            f.remediation.effort_estimate, f.remediation.priority,
            "Required" if f.remediation.external_validation_required else "Not required",
            f.remediation.validation_reason,
            f.remediation.validation_query,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font()
            c.alignment = _align(wrap=(col in (4, 5, 6, 7, 12, 13)))

        if i % 2 == 1:
            for col in range(1, len(COLS) + 1):
                ws.cell(row=r, column=col).fill = _fill(_C_ALT)

        sev = f.severity
        ws.cell(row=r, column=2).fill = _fill(_SEV_BG.get(sev, "FFFFFF"))
        ws.cell(row=r, column=2).font = _font(bold=True, color=_SEV_FG.get(sev, "000000"))
        ws.cell(row=r, column=2).alignment = _align(h="center")

    _autofit_columns(ws, min_width=8, max_width=55)


def _build_cross_references(wb: Workbook, xrefs: list[CrossReference]):
    ws = wb.create_sheet("Cross-References")
    _apply_app_style(ws, zoom=110)
    ws.sheet_properties.tabColor = _TAB_XREF

    COLS = ["Title", "Related Findings", "Correlation", "Combined Impact", "Recommended Action"]
    _hdr_row(ws, 1, COLS)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(5)}1"

    for i, xr in enumerate(xrefs):
        r = 2 + i
        vals = [
            xr.title,
            ", ".join(xr.related_findings),
            xr.correlation_description,
            xr.combined_impact,
            xr.recommended_action,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font()
            c.alignment = _align(wrap=True)
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = _fill(_C_ALT)

    _autofit_columns(ws, min_width=10, max_width=55)


def _build_remediation_plan(wb: Workbook, analyses: list[DomainAnalysis]):
    ws = wb.create_sheet("Remediation Plan")
    _apply_app_style(ws, zoom=110)
    ws.sheet_properties.tabColor = _TAB_REM

    COLS = [
        "Finding ID", "Section", "Severity", "Title", "Action", "Transactions",
        "Effort", "Priority", "External Validation", "Validation Reason",
        "Validation Query",
    ]
    _hdr_row(ws, 1, COLS)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    _SEV_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    rows: list[tuple] = []
    for da in analyses:
        for f in da.findings:
            rows.append((
                _SEV_ORDER.get(f.severity, 9),
                f.id, da.section_title, f.severity, f.title,
                f.remediation.action, ", ".join(f.remediation.sap_transactions),
                f.remediation.effort_estimate, f.remediation.priority,
                "Required" if f.remediation.external_validation_required else "Not required",
                f.remediation.validation_reason,
                f.remediation.validation_query,
            ))
    rows.sort(key=lambda x: x[0])

    for i, row_data in enumerate(rows):
        r = 2 + i
        vals = row_data[1:]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font()
            c.alignment = _align(wrap=(col in (5, 6, 10, 11)))

        if i % 2 == 1:
            for col in range(1, len(COLS) + 1):
                ws.cell(r, col).fill = _fill(_C_ALT)

        sev = vals[2]
        ws.cell(r, 3).fill = _fill(_SEV_BG.get(sev, "FFFFFF"))
        ws.cell(r, 3).font = _font(bold=True, color=_SEV_FG.get(sev, "000000"))
        ws.cell(r, 3).alignment = _align(h="center")

    _autofit_columns(ws, min_width=8, max_width=55)


def _build_document_structure(wb: Workbook, tree: dict):
    ws = wb.create_sheet("Document Structure")
    _apply_app_style(ws, zoom=110)
    ws.sheet_properties.tabColor = _TAB_UTIL

    _hdr_row(ws, 1, ["Title", "Pages", "Summary", "Depth"])
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    row = [2]

    def _walk(nodes, depth=0):
        for n in nodes:
            title = ("  " * depth) + n.get("title", "")
            ps, pe = n.get("start_index", ""), n.get("end_index", "")
            pages = f"{ps}–{pe}" if ps and pe else ""
            summary = n.get("summary", "")
            r = row[0]
            ws.cell(r, 1, title).font = _font(bold=(depth == 0))
            ws.cell(r, 2, pages).alignment = _align(h="center")
            c = ws.cell(r, 3, summary)
            c.font = _font()
            c.alignment = _align(wrap=True)
            ws.cell(r, 4, depth).font = _font()
            if r % 2 == 0:
                for col in range(1, 5):
                    ws.cell(r, col).fill = _fill(_C_ALT)
            row[0] += 1
            _walk(n.get("nodes", []), depth + 1)

    _walk(tree.get("structure", []))
    _autofit_columns(ws, min_width=8, max_width=55)


def _build_token_usage(wb: Workbook, usage: TokenUsage):
    ws = wb.create_sheet("Token Usage")
    _apply_app_style(ws, zoom=110)
    ws.sheet_properties.tabColor = _TAB_UTIL

    _hdr_row(ws, 1, ["Phase", "Model", "Input Tokens", "Cached Input", "Cache Write", "Billable Input", "Output Tokens", "Total Tokens"])
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    data_rows = [
        ("Phase 0 — Planning", usage.phase0_model, usage.phase0_input_tokens, usage.phase0_cached_input_tokens, usage.phase0_cache_write_tokens, usage.phase0_output_tokens),
        ("Phase 1 — Domain Analysis", usage.phase1_model, usage.phase1_input_tokens, usage.phase1_cached_input_tokens, usage.phase1_cache_write_tokens, usage.phase1_output_tokens),
        ("Phase 2 — Cross-Ref+Synth", usage.phase2_model, usage.phase2_input_tokens, usage.phase2_cached_input_tokens, usage.phase2_cache_write_tokens, usage.phase2_output_tokens),
    ]

    for i, (phase, model, inp, cached_inp, cache_write, out) in enumerate(data_rows):
        r = 2 + i
        ws.cell(r, 1, phase).font = _font()
        ws.cell(r, 2, model).font = _font()
        ws.cell(r, 3, inp).number_format = _NUM
        ws.cell(r, 4, cached_inp).number_format = _NUM
        ws.cell(r, 5, cache_write).number_format = _NUM
        ws.cell(r, 6, max(inp - cached_inp - cache_write, 0)).number_format = _NUM
        ws.cell(r, 7, out).number_format = _NUM
        c = ws.cell(r, 8, f"=C{r}+G{r}")
        c.number_format = _NUM
        c.font = _font()
        if i % 2 == 1:
            for col in range(1, 8):
                ws.cell(r, col).fill = _fill(_C_ALT)

    tr = 2 + len(data_rows)
    ws.cell(tr, 1, "TOTAL").font = _font(bold=True)
    for col, letter in ((3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")):
        c = ws.cell(tr, col, f"=SUM({letter}2:{letter}{tr-1})")
        c.number_format = _NUM
        c.font = _font(bold=True)
    for col in range(1, 8):
        ws.cell(tr, col).fill = _fill("BDD7EE")

    ws.cell(tr + 2, 1,
        "Phase 0 tokens are not tracked — PageIndex runs via LiteLLM independently."
    ).font = _font(italic=True, color="888888")

    _autofit_columns(ws, min_width=8, max_width=40)


# ── Public entry point ────────────────────────────────────────────────────────

def generate(result: AnalysisResult, output_path: Path, tree: dict | None = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _ILLEGAL_SHEET = str.maketrans("", "", "/\\?*[]:'")

    sheet_map: dict[str, str] = {}
    used: set[str] = set()
    for da in result.domain_analyses:
        if not da.findings:
            continue
        clean = da.section_title.translate(_ILLEGAL_SHEET).strip()[:31] or "Section"
        base, suffix, name = clean, 1, clean
        while name in used:
            name = f"{base[:28]}_{suffix}"
            suffix += 1
        sheet_map[da.section_id] = name
        used.add(name)

    _build_executive_summary(wb, result, sheet_map)
    for da in result.domain_analyses:
        if da.section_id in sheet_map:
            _build_section_sheet(wb, da, sheet_map[da.section_id])
    if result.cross_references:
        _build_cross_references(wb, result.cross_references)
    _build_remediation_plan(wb, result.domain_analyses)
    if tree:
        _build_document_structure(wb, tree)
    _build_token_usage(wb, result.token_usage)

    # Open on the dashboard
    wb.active = 0

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
